import re
from typing import Optional
from openpyxl import load_workbook

#RFC MEXICO: 3-4 LETRAS (&,Ñ) + 6 Digitos (fecha) + 3 alfanum
RFC_REGEX = re.compile(r"\b([A-Z&Ñ]{3,4}\d{6}[A-Z0-9]{3})\b", re.IGNORECASE)

def extract_rfc_from_xlsx(file_obj) -> Optional[str]:
    """
    Extracts the first RFC found in an XLSX file-like object.
    Returns RFC uppercase or None if not found.
    """
    wb = load_workbook(filename = file_obj, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell)
                    m = RFC_REGEX.search(text.upper())
                    if m:
                        return m.group(1).upper()
        return None
    finally:
        wb.close()
