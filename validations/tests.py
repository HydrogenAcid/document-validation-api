from io import BytesIO
from django.test import TestCase
from openpyxl import Workbook

from .extractor import extract_rfc_from_xlsx

class ExtractorTests(TestCase):
    def test_extract_rfc_from_xlsx_success(self):
        """
        Requirement: extractor returns expected value when RFC exists.
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "RFC"
        ws["B1"] = "PEPJ8001019Q8"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        rfc = extract_rfc_from_xlsx(buf)
        self.assertEqual(rfc, "PEPJ8001019Q8")

    def test_extract_rfc_from_xlsx_not_found(self):
        """
        Extractor should fail in a controlled way (return None) when no RFC exists.
        """
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "NO RFC HERE"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        rfc = extract_rfc_from_xlsx(buf)
        self.assertIsNone(rfc)